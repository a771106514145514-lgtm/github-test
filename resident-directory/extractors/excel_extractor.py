"""Excel (.xlsx/.xls) 擷取器。

策略：先嘗試辨識標題列（姓名/電話/地址/戶號等關鍵字），若成功則逐列直接對應欄位，
準確度遠高於純文字規則比對；若找不到可辨識的標題列，則整列文字存成 raw_text 並標記需複核。
"""
import openpyxl

from normalize import match_excel_header, extract_fields_from_text


def _find_header_row(sheet, max_scan_rows: int = 5):
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan_rows), start=1):
        mapping = {}
        for col_idx, cell in enumerate(row):
            field = match_excel_header(cell.value)
            if field and field not in mapping:
                mapping[col_idx] = field
        if len(mapping) >= 2:
            return row_idx, mapping
    return None, {}


def extract_excel(file_path: str) -> list:
    records = []
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    for sheet in workbook.worksheets:
        header_row_idx, mapping = _find_header_row(sheet)

        if mapping:
            for row in sheet.iter_rows(min_row=header_row_idx + 1):
                values = [cell.value for cell in row]
                if not any(values):
                    continue
                fields = {'name': '', 'phone': '', 'address': '', 'unit': ''}
                for col_idx, field in mapping.items():
                    if col_idx < len(values) and values[col_idx] is not None:
                        fields[field] = str(values[col_idx]).strip()
                if not (fields['name'] or fields['phone']):
                    continue
                raw_text = ' '.join(str(v) for v in values if v is not None)
                found = sum(1 for v in fields.values() if v)
                confidence = 'high' if found >= 2 else 'medium'
                records.append({
                    **fields,
                    'raw_text': raw_text,
                    'confidence': confidence,
                    'needs_review': confidence != 'high',
                    'sheet_name': sheet.title,
                })
        else:
            # 找不到清楚的標題列，逐列丟給文字規則比對，找不到欄位的整列標記需複核
            for row in sheet.iter_rows():
                values = [cell.value for cell in row if cell.value is not None]
                if not values:
                    continue
                raw_text = ' '.join(str(v) for v in values)
                fields = extract_fields_from_text(raw_text)
                records.append({**fields, 'raw_text': raw_text, 'sheet_name': sheet.title})

    workbook.close()
    return records
